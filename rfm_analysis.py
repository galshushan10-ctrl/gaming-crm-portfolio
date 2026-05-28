import pandas as pd
import numpy as np
from datetime import datetime, timedelta
import random
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference
from openpyxl.chart.label import DataLabelList

# ─── CONFIG ───────────────────────────────────────────────────────────────────
TODAY = datetime(2026, 5, 28)
N_PLAYERS = 1000
RANDOM_SEED = 42
OUTPUT_FILE = "gaming_rfm_report.xlsx"

# ─── SEGMENT DEFINITIONS ──────────────────────────────────────────────────────
SEGMENT_CONFIG = {
    "Whale / VIP":          {"color": "FFD700", "priority": 1},
    "Champion":             {"color": "00B050", "priority": 2},
    "Loyal Player":         {"color": "92D050", "priority": 3},
    "New Player":           {"color": "00B0F0", "priority": 4},
    "Active Free Player":   {"color": "0070C0", "priority": 5},
    "At Risk":              {"color": "FF9900", "priority": 6},
    "High Value Churning":  {"color": "FF0000", "priority": 7},
    "Churned":              {"color": "C00000", "priority": 8},
    "Casual":               {"color": "BFBFBF", "priority": 9},
    "Dormant":              {"color": "808080", "priority": 10},
}

SEGMENT_ACTIONS = {
    "Whale / VIP":         "VIP program, exclusive rewards, personal account manager",
    "Champion":            "Beta access, loyalty rewards, upsell premium features",
    "Loyal Player":        "Cross-sell, referral program, seasonal bonuses",
    "New Player":          "Onboarding flow, tutorial bonus, first-purchase offer",
    "Active Free Player":  "Starter pack offers, limited-time deals, monetization push",
    "At Risk":             "Win-back campaign, special offer, re-engagement push notification",
    "High Value Churning": "URGENT: Personal outreach, high-value retention offer",
    "Churned":             "Win-back email series, comeback bonus (30/60/90 day)",
    "Casual":              "Habit-building nudges, daily reward streaks",
    "Dormant":             "Last-chance re-engagement or sunset",
}


# ─── DATA GENERATION ──────────────────────────────────────────────────────────
def generate_data(n: int) -> pd.DataFrame:
    np.random.seed(RANDOM_SEED)
    random.seed(RANDOM_SEED)

    profiles = [
        # (label, weight, recency_range, freq_range, spend_range)
        ("whale",    0.05, (1,  14), (20, 60), (200, 2000)),
        ("regular",  0.25, (1,  30), ( 8, 25), ( 20,  200)),
        ("casual",   0.30, (7,  60), ( 2, 10), (  0,   50)),
        ("at_risk",  0.20, (30, 90), ( 3, 15), ( 10,  100)),
        ("churned",  0.20, (90,365), ( 1,  5), (  0,   30)),
    ]
    weights    = [p[1] for p in profiles]
    games      = ["Battle Arena", "Fantasy Quest", "Speed Racers", "Space Wars", "Puzzle Kingdom"]
    platforms  = ["Mobile", "PC", "Console"]
    p_weights  = [0.50, 0.30, 0.20]
    countries  = ["Israel", "USA", "UK", "Germany", "France", "Brazil", "Japan", "Korea"]
    c_weights  = [0.15, 0.25, 0.10, 0.08, 0.07, 0.10, 0.10, 0.15]

    rows = []
    for i in range(n):
        _, _, rec_r, freq_r, spend_r = random.choices(profiles, weights=weights)[0]

        recency   = random.randint(*rec_r)
        frequency = random.randint(*freq_r)
        monetary  = round(random.uniform(*spend_r), 2)
        reg_days  = random.randint(max(recency, 30), 1095)

        rows.append({
            "player_id":               f"PLR_{i+1:04d}",
            "registration_date":       (TODAY - timedelta(days=reg_days)).strftime("%Y-%m-%d"),
            "last_activity_date":      (TODAY - timedelta(days=recency)).strftime("%Y-%m-%d"),
            "days_since_last_activity": recency,
            "sessions_last_90d":       frequency,
            "total_spend_usd":         monetary,
            "favorite_game":           random.choice(games),
            "platform":                random.choices(platforms, weights=p_weights)[0],
            "country":                 random.choices(countries, weights=c_weights)[0],
            "player_level":            min(100, int(frequency * random.uniform(0.8, 1.5))),
        })

    return pd.DataFrame(rows)


# ─── RFM SCORING ──────────────────────────────────────────────────────────────
def score_rfm(df: pd.DataFrame) -> pd.DataFrame:
    df = df.copy()

    df["R_score"] = pd.qcut(
        df["days_since_last_activity"], q=5, labels=[5, 4, 3, 2, 1]
    ).astype(int)

    df["F_score"] = pd.qcut(
        df["sessions_last_90d"].rank(method="first"), q=5, labels=[1, 2, 3, 4, 5]
    ).astype(int)

    df["M_score"] = pd.qcut(
        df["total_spend_usd"].rank(method="first"), q=5, labels=[1, 2, 3, 4, 5]
    ).astype(int)

    df["RFM_score"] = (
        df["R_score"].astype(str) + df["F_score"].astype(str) + df["M_score"].astype(str)
    )
    df["RFM_total"] = df["R_score"] + df["F_score"] + df["M_score"]

    return df


# ─── SEGMENTATION ─────────────────────────────────────────────────────────────
def assign_segment(row) -> str:
    r, f, m = row["R_score"], row["F_score"], row["M_score"]

    if r >= 4 and f >= 4 and m >= 4:
        return "Whale / VIP"
    if r >= 4 and f >= 3 and m >= 3:
        return "Champion"
    if r >= 3 and f >= 3 and m >= 3:
        return "Loyal Player"
    if r >= 4 and f <= 2:
        return "New Player"
    if r >= 3 and f >= 3 and m <= 2:
        return "Active Free Player"
    if r <= 2 and f >= 3 and m >= 3:
        return "High Value Churning"
    if r <= 2 and f >= 3:
        return "At Risk"
    if r == 1 and f <= 2:
        return "Churned"
    if r >= 3 and f <= 2:
        return "Casual"
    return "Dormant"


# ─── EXCEL STYLING HELPERS ────────────────────────────────────────────────────
def fill(hex_color: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_color)

def bold_font(size=11, color="000000", white=False) -> Font:
    return Font(bold=True, size=size, color="FFFFFF" if white else color)

def center() -> Alignment:
    return Alignment(horizontal="center", vertical="center", wrap_text=True)

def thin_border() -> Border:
    s = Side(style="thin", color="D9D9D9")
    return Border(left=s, right=s, top=s, bottom=s)

HEADER_FILL  = fill("1F3864")
HEADER_FONT  = bold_font(11, white=True)
STRIPE_FILL  = fill("F2F2F2")


def style_header_row(ws, row: int, n_cols: int):
    for col in range(1, n_cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill   = HEADER_FILL
        cell.font   = HEADER_FONT
        cell.alignment = center()
        cell.border = thin_border()


def autofit_columns(ws, min_width=10, max_width=40):
    for col in ws.columns:
        col_letter = get_column_letter(col[0].column)
        lengths = [len(str(c.value)) if c.value else 0 for c in col]
        width = min(max(max(lengths) + 2, min_width), max_width)
        ws.column_dimensions[col_letter].width = width


# ─── SHEET 1: RAW DATA ────────────────────────────────────────────────────────
def write_raw_sheet(wb: openpyxl.Workbook, df: pd.DataFrame):
    ws = wb.active
    ws.title = "Player Data"
    ws.freeze_panes = "A2"

    headers = list(df.columns)
    ws.append(headers)
    style_header_row(ws, 1, len(headers))

    seg_col = headers.index("segment") + 1 if "segment" in headers else None

    for i, row in enumerate(df.itertuples(index=False), start=2):
        ws.append(list(row))
        stripe = i % 2 == 0
        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=i, column=col)
            cell.border    = thin_border()
            cell.alignment = Alignment(vertical="center")
            if stripe and (seg_col is None or col != seg_col):
                cell.fill = STRIPE_FILL

        if seg_col:
            seg_cell = ws.cell(row=i, column=seg_col)
            seg_name = seg_cell.value
            if seg_name in SEGMENT_CONFIG:
                seg_cell.fill = fill(SEGMENT_CONFIG[seg_name]["color"])
                seg_cell.font = Font(bold=True, color="FFFFFF" if seg_name in
                    {"Whale / VIP", "Churned", "High Value Churning", "At Risk",
                     "Active Free Player", "Dormant"} else "000000")

    ws.row_dimensions[1].height = 30
    autofit_columns(ws)


# ─── SHEET 2: SEGMENT SUMMARY ─────────────────────────────────────────────────
def write_summary_sheet(wb: openpyxl.Workbook, df: pd.DataFrame):
    ws = wb.create_sheet("Segment Summary")
    ws.freeze_panes = "A2"

    summary = (
        df.groupby("segment")
        .agg(
            player_count=("player_id", "count"),
            avg_recency=("days_since_last_activity", "mean"),
            avg_sessions=("sessions_last_90d", "mean"),
            avg_spend=("total_spend_usd", "mean"),
            total_revenue=("total_spend_usd", "sum"),
        )
        .reset_index()
    )
    summary["pct_of_base"]    = (summary["player_count"] / len(df) * 100).round(1)
    summary["avg_recency"]    = summary["avg_recency"].round(1)
    summary["avg_sessions"]   = summary["avg_sessions"].round(1)
    summary["avg_spend"]      = summary["avg_spend"].round(2)
    summary["total_revenue"]  = summary["total_revenue"].round(2)
    summary["recommended_action"] = summary["segment"].map(SEGMENT_ACTIONS)
    summary["priority"]       = summary["segment"].map(
        lambda s: SEGMENT_CONFIG.get(s, {}).get("priority", 99)
    )
    summary = summary.sort_values("priority").drop(columns="priority")

    headers = [
        "Segment", "# Players", "% of Base", "Avg Days Since Login",
        "Avg Sessions (90d)", "Avg Spend ($)", "Total Revenue ($)", "Recommended Action"
    ]
    ws.append(headers)
    style_header_row(ws, 1, len(headers))
    ws.row_dimensions[1].height = 30

    for i, row in enumerate(summary.itertuples(index=False), start=2):
        ws.append(list(row))
        seg_name = row.segment
        cfg      = SEGMENT_CONFIG.get(seg_name, {})
        hex_col  = cfg.get("color", "FFFFFF")
        dark_bg  = seg_name in {"Whale / VIP", "Churned", "High Value Churning",
                                "At Risk", "Active Free Player", "Dormant"}

        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=i, column=col)
            cell.border    = thin_border()
            cell.alignment = Alignment(vertical="center", wrap_text=(col == len(headers)))
            if col == 1:
                cell.fill = fill(hex_col)
                cell.font = Font(bold=True, color="FFFFFF" if dark_bg else "000000")
            elif i % 2 == 0:
                cell.fill = STRIPE_FILL

    ws.row_dimensions[1].height = 30
    for i in range(2, len(summary) + 2):
        ws.row_dimensions[i].height = 25
    autofit_columns(ws, min_width=12, max_width=50)


# ─── SHEET 3: SEGMENT CHART ───────────────────────────────────────────────────
def write_chart_sheet(wb: openpyxl.Workbook, df: pd.DataFrame):
    ws = wb.create_sheet("Charts")

    seg_counts = (
        df["segment"]
        .value_counts()
        .reindex([s for s in SEGMENT_CONFIG if s in df["segment"].unique()])
        .reset_index()
    )
    seg_counts.columns = ["Segment", "Count"]

    # Write mini-table for chart
    ws["A1"] = "Segment"
    ws["B1"] = "Players"
    style_header_row(ws, 1, 2)

    for i, row in enumerate(seg_counts.itertuples(index=False), start=2):
        ws.cell(row=i, column=1, value=row.Segment)
        ws.cell(row=i, column=2, value=row.Count)

    # Bar chart
    chart = BarChart()
    chart.type   = "col"
    chart.title  = "Players per Segment"
    chart.y_axis.title = "# Players"
    chart.x_axis.title = "Segment"
    chart.style  = 10
    chart.width  = 22
    chart.height = 14

    data = Reference(ws, min_col=2, min_row=1, max_row=len(seg_counts) + 1)
    cats = Reference(ws, min_col=1, min_row=2, max_row=len(seg_counts) + 1)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.dataLabels = DataLabelList()
    chart.dataLabels.showVal = True

    ws.add_chart(chart, "D2")
    autofit_columns(ws)


# ─── SHEET 4: KPI DASHBOARD ───────────────────────────────────────────────────
def write_kpi_sheet(wb: openpyxl.Workbook, df: pd.DataFrame):
    ws = wb.create_sheet("KPI Dashboard")

    total     = len(df)
    total_rev = df["total_spend_usd"].sum()
    vip       = df[df["segment"] == "Whale / VIP"]
    churned   = df[df["segment"].isin(["Churned", "Dormant"])]
    at_risk   = df[df["segment"].isin(["At Risk", "High Value Churning"])]
    paying    = df[df["total_spend_usd"] > 0]

    kpis = [
        ("Total Players",             total,                                     ""),
        ("Total Revenue ($)",          round(total_rev, 2),                      ""),
        ("Avg Revenue per Player ($)", round(total_rev / total, 2),              ""),
        ("Whales / VIPs",             len(vip),                                  f"{len(vip)/total*100:.1f}% of base"),
        ("Whale Revenue ($)",          round(vip["total_spend_usd"].sum(), 2),   f"{vip['total_spend_usd'].sum()/total_rev*100:.1f}% of revenue"),
        ("Paying Players",            len(paying),                               f"{len(paying)/total*100:.1f}% conversion"),
        ("At-Risk Players",           len(at_risk),                              f"{len(at_risk)/total*100:.1f}% of base — take action"),
        ("Churned + Dormant",         len(churned),                              f"{len(churned)/total*100:.1f}% of base"),
        ("Avg Days Since Login",       round(df["days_since_last_activity"].mean(), 1), "days"),
        ("Avg Sessions (90d)",         round(df["sessions_last_90d"].mean(), 1), "sessions"),
    ]

    ws["A1"] = "KPI"
    ws["B1"] = "Value"
    ws["C1"] = "Notes"
    style_header_row(ws, 1, 3)
    ws.row_dimensions[1].height = 25

    for i, (kpi, val, note) in enumerate(kpis, start=2):
        ws.cell(row=i, column=1, value=kpi).font  = Font(bold=True)
        ws.cell(row=i, column=2, value=val).alignment = Alignment(horizontal="center")
        ws.cell(row=i, column=3, value=note)
        for col in range(1, 4):
            c = ws.cell(row=i, column=col)
            c.border    = thin_border()
            c.alignment = Alignment(vertical="center")
            if i % 2 == 0:
                c.fill = STRIPE_FILL
        ws.row_dimensions[i].height = 22

    autofit_columns(ws, min_width=20, max_width=50)


# ─── MAIN ─────────────────────────────────────────────────────────────────────
def main():
    print("Generating player data...")
    df = generate_data(N_PLAYERS)

    print("Calculating RFM scores...")
    df = score_rfm(df)
    df["segment"]            = df.apply(assign_segment, axis=1)
    df["recommended_action"] = df["segment"].map(SEGMENT_ACTIONS)

    print("Building Excel report...")
    wb = openpyxl.Workbook()
    write_raw_sheet(wb, df)
    write_summary_sheet(wb, df)
    write_chart_sheet(wb, df)
    write_kpi_sheet(wb, df)

    wb.save(OUTPUT_FILE)
    print(f"\nDone! Report saved to: {OUTPUT_FILE}")
    print(f"\nSegment breakdown:")
    print(df["segment"].value_counts().to_string())


if __name__ == "__main__":
    main()
